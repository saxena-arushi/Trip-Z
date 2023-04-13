from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User, auth
from django.contrib import messages
from .models import UserDetails, Preferences, PublishedTrips
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.http import HttpResponse
import openpyxl
# Create your views here.

def LandingPage(request):
    return render(request,"index.html")
	
def signup(request):

    if request.method == 'POST':
        firstName = request.POST['fname']
        lastName = request.POST['lname']
        email = request.POST['email']
        username = request.POST['username']
        password = request.POST['pass']
        cpassword = request.POST['passconfirm']
        pno = request.POST['contact']
        dob = request.POST['dob']
        city = request.POST['city']
        gender = request.POST['gender']
        securityQues = request.POST['secques']
        securityAns = request.POST['secans']
        bio = request.POST['bio']

        if password == cpassword:
            if User.objects.filter(email = email).exists():
                messages.info(request, 'Email already exists!')
                return redirect('signup')
            elif User.objects.filter(username = username).exists():
                messages.info(request, 'User already exists!')
                return redirect('signup')
            else:
                user = User.objects.create_user(username = username , email = email, password = password, first_name=firstName, last_name=lastName)
                user.save()

            user = authenticate(request, username=username, password=password)
            login(request, user)

            user = request.user
            u = UserDetails.objects.create(
                user=user, 
                pno=pno, 
                city=city, 
                dob=dob, 
                gender=gender, 
                securityQues=securityQues, 
                securityAns=securityAns, 
                bio=bio
            )
  
            u.save()

                
            return redirect('preferences')         
    return render(request,"signup.html")

def Login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['pass']
        secAns = request.POST['secAns']

        user = authenticate(username = username, password = password)

        if user is not None and UserDetails.objects.filter(securityAns = secAns).exists():
            auth.login(request, user)
            return redirect('recommendations')
        else:
            messages.info(request, 'Credentials Invalid!')
            return redirect('Login') 
    return render(request,"login.html")

def logout(request):
    auth.logout(request)
    return redirect('/')

@login_required
def preferences(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            user = request.user
            age = request.POST.getlist('age[]')
            gender = request.POST.getlist('gender[]')
            maritalStatus = request.POST.getlist('maritalStatus[]')
            travelFrequency = request.POST.getlist('travelFrequency[]')
            tripType = request.POST.getlist('tripType[]')
            budget = request.POST.getlist('budget[]')
            travelCompanion = request.POST.getlist('travelCompanion[]')
            communicationStyle = request.POST.getlist('communicationStyle[]')
            interests = request.POST.getlist('interests[]')
            kindOfTrips = request.POST.getlist('kindOfTrips[]')
            travelDestination = request.POST.getlist('travelDestination[]')
            stays = request.POST.getlist('stays[]')
            travelMotivation = request.POST.getlist('travelMotivation[]')
            requirement = request.POST.getlist('requirement[]')
            modeOfTravel = request.POST.getlist('modeOfTravel[]')
            travelPace = request.POST.getlist('travelPace[]')
            eatingPlaces = request.POST.getlist('eatingPlaces[]')
            accomodations = request.POST.getlist('accomodations[]')

            p = Preferences.objects.create(
                user = user,
                age = age,
                gender = gender,
                maritalStatus = maritalStatus,
                travelFrequency = travelFrequency,
                tripType = tripType,
                budget = budget,
                travelCompanion = travelCompanion,
                communicationStyle = communicationStyle,
                interests = interests,
                kindOfTrips = kindOfTrips,
                travelDestination = travelDestination,
                stays = stays,
                travelMotivation = travelMotivation,
                requirement = requirement,
                modeOfTravel = modeOfTravel,
                travelPace = travelPace,
                eatingPlaces = eatingPlaces,
                accomodations = accomodations
            )
            p.save()

            # # add user data to Excel file
            # try:
            #     workbook = openpyxl.load_workbook('UserTrips//users.xlsx')
            # except FileNotFoundError:
            #     workbook = openpyxl.Workbook()
            # worksheet = workbook.active
            # row_number = worksheet.max_row + 1

            # # age_string = ",".join(age)
            # # gender_string = ",".join(gender)
            # # maritalStatus_string = ",".join(maritalStatus)
            # # travelFrequency_string = ",".join(travelFrequency)
            # # tripType_string = ",".join(tripType)
            # # budget_string = ",".join(budget)
            # # travelCompanion_string = ",".join(travelCompanion)
            # # communicationStyle_string = ",".join(communicationStyle)
            # # interests_string = ",".join(interests)
            # # kindOfTrips_string = ",".join(kindOfTrips)
            # # travelDestination_string = ",".join(travelDestination)
            # # stays_string = ",".join(stays)
            # # travelMotivation_string = ",".join(travelMotivation)
            # # requirement_string = ",".join(requirement)
            # # modeOfTravel_string = ",".join(modeOfTravel)
            # # travelPace_string = ",".join(travelPace)
            # # eatingPlaces_string = ",".join(eatingPlaces)
            # # accomodations_string = ",".join(accomodations)

            # # data = [
            # #     age,
            # #     gender,
            # #     maritalStatus,
            # #     travelFrequency,
            # #     tripType,
            # #     budget,
            # #     travelCompanion,
            # #     communicationStyle,
            # #     interests,
            # #     kindOfTrips,
            # #     travelDestination,
            # #     stays,
            # #     travelMotivation,
            # #     requirement,
            # #     modeOfTravel,
            # #     travelPace,
            # #     eatingPlaces,
            # #     accomodations
            # # ]
            # # add user data to Excel file
            # try:
            #     workbook = openpyxl.load_workbook('UserTrips//users.xlsx')
            # except FileNotFoundError:
            #     workbook = openpyxl.Workbook()
            # worksheet = workbook.active
            # row_number = worksheet.max_row + 1
        
            # # # add the data to the specific columns in the Excel file
            # # worksheet.cell(row=row_number, column=1).value = ','.join(age)
            # # worksheet.cell(row=row_number, column=2).value = ','.join(gender)
            # # worksheet.cell(row=row_number, column=3).value = ','.join(maritalStatus)
            # # worksheet.cell(row=row_number, column=4).value = ','.join(travelFrequency)
            # # worksheet.cell(row=row_number, column=5).value = ','.join(tripType)
            # # worksheet.cell(row=row_number, column=6).value = ','.join(budget)
            # # worksheet.cell(row=row_number, column=7).value =
            # # workbook.save('users.xlsx')
            # worksheet.cell(row=row_number, column=1).value = age[0]
            # worksheet.cell(row=row_number, column=2).value = gender[0]
            # worksheet.cell(row=row_number, column=3).value = maritalStatus[0]
            # worksheet.cell(row=row_number, column=4).value = travelFrequency[0]
            # worksheet.cell(row=row_number, column=5).value = tripType[0]
            # worksheet.cell(row=row_number, column=6).value = budget[0]
            # worksheet.cell(row=row_number, column=7).value = travelCompanion[0]
            # worksheet.cell(row=row_number, column=8).value = communicationStyle[0]
            # worksheet.cell(row=row_number, column=9).value = interests[0]
            # worksheet.cell(row=row_number, column=10).value = kindOfTrips[0]
            # worksheet.cell(row=row_number, column=11).value = travelDestination[0]
            # worksheet.cell(row=row_number, column=12).value = stays[0]
            # worksheet.cell(row=row_number, column=13).value = travelMotivation[0]
            # worksheet.cell(row=row_number, column=14).value = requirement[0]
            # worksheet.cell(row=row_number, column=15).value = modeOfTravel[0]
            # worksheet.cell(row=row_number, column=16).value = travelPace[0]
            # worksheet.cell(row=row_number, column=17).value = eatingPlaces[0]
            # worksheet.cell(row=row_number, column=18).value = accomodations[0]

            # # for row_data in data:
            # #     worksheet.append(row_data)
            
            # workbook.save('users.xlsx')
            
            return redirect('recommendations')
    return render(request, 'preferences.html')


def recommendations(request):
    if request.user.is_authenticated:
        #retrieve(request)
        if request.method == 'GET':
            print(request.GET)
            leavingFrom = request.GET.get('goingFrom')
            goingTo = request.GET.get('goingTo')
            startDate = request.GET.get('startDate')
            endDate = request.GET.get('endDate')

            trips = PublishedTrips.objects.filter(
                leavingFrom=leavingFrom,
                goingTo=goingTo,
                startDate=startDate,
                endDate=endDate,
            )
            print(trips)
            if trips.exists():
                users = User.objects.filter(publishedtrips__in=trips).distinct()
                # users = User.objects.filter(publishedtrips__in=trips).distinct().values_list('username', flat=True)
                user_details_list = []
                for user in users:
                    user_details = user.userdetails.get()
                    user_details_list.append(user_details)
                return render(request, 'recommendations.html', {'users': users, 'user_details' : user_details_list })
            else:
                return render(request, 'recommendations.html', {'users': []})
        return render(request, 'recommendations.html')
    else:
        return redirect('/')

from django.db.models import Q

@login_required
def retrieve(request):
    users = User.objects.filter(email__icontains='@gmail.com')
    #user_details_list = []
    # for user in users:
    #     user_details = user.userdetails.get()
    #     user_details_list.append(user_details)
    print(users)
    #print(user_details_list)        
    return render(request, 'retrieve.html', {'users': users,})
    
                #users = User.objects.filter(publishedtrips__in=trips).distinct()

    
# get the user whose marital status we want to match
    #user = User.objects.get(username = request.user.username)

# get all users with the same marital status preference
    user_dict = {}

    for user in users:
        marital_status = user.marital_status
        if marital_status not in user_dict:
            user_dict[marital_status] = []
        user_dict[marital_status].append(user)

    print(user_dict)
#This will print a dictionary where the keys are the different marital status values, and the values are lists of users with the corresponding marital status. You can then pass this dictionary to your HTML template to display the users grouped by their marital status.






@login_required
def personalDashboard(request):
    user = request.user
    userD = UserDetails.objects.get(user = user)
    try : 
        publishedTrips = PublishedTrips.objects.filter(user = user)
        # c = publishedTrips.count()
        return render(request, 'personal-dashboard.html', { 'user': user, 'userD': userD, 'publishedTrips': publishedTrips, })

    except PublishedTrips.DoesNotExist:
        publishedTrips = None
        return render(request, 'personal-dashboard.html', { 'user': user, 'userD': userD })

@login_required
def publish_a_trip(request):
    if request.method == 'POST':
        user = request.user
        goingFrom = request.POST['goingFrom']
        goingTo = request.POST.get('goingTo')
        startDate = request.POST.get('startDate')
        endDate = request.POST.get('endDate')
        noOfPersons = request.POST.get('noOfPeople')

        p = PublishedTrips.objects.create(
            user=user,
            leavingFrom=goingFrom,
            goingTo=goingTo,
            startDate=startDate,
            endDate=endDate,
            noOfPersons = noOfPersons
        )
        p.save()
        return redirect('/personal-dashboard')
    return render(request, 'publish_trip.html')

# def publish_trip(request):
#     return render(request, 'publish_trip.html')
#publish_a_trip(request)

   
def aboutus(request):
    return render(request,"aboutus.html")

def contactus(request):
    return render(request,"contactus.html")

def userprofile(request, username):
    if request.method == 'GET':
        user = get_object_or_404(User, username=username)
        user_details = UserDetails.objects.get(user = user)
        trips = PublishedTrips.objects.filter(user = user)
        return render(request, "userprofile.html", {'user' : user, 'user_details': user_details, 'trips' : trips})
    return render(request, "userprofile.html")

